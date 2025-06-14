import json
import os
from typing import List, Dict, Any

import openpyxl

from .base import VacancyStorage
from ..models import Vacancy


class ExcelVacancyStorage(VacancyStorage):
    """Класс для сохранения вакансий в Excel-файл."""

    def __init__(self, file_path: str):
        self.file_path = file_path
        self._ensure_file_exists()

    def _ensure_file_exists(self) -> None:
        try:
            workbook = openpyxl.load_workbook(self.file_path)
            workbook.close()
        except FileNotFoundError:
            # Создаем директорию, если она не существует
            import os
            os.makedirs(os.path.dirname(self.file_path), exist_ok=True)

            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.append(["title", "link", "salary", "description", "requirements"])
            workbook.save(self.file_path)
            workbook.close()
        except Exception as e:
            print(f"Ошибка при создании файла {self.file_path}: {e}")

    def add_vacancy(self, vacancy: Vacancy) -> None:
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        sheet.append([
            vacancy.title,
            vacancy.link,
            json.dumps(vacancy.salary) if vacancy.salary else "",
            vacancy.description,
            vacancy.requirements,
        ])
        workbook.save(self.file_path)
        workbook.close()

    def get_vacancies(self, criteria: Dict[str, Any]) -> List[Vacancy]:
        vacancies = []
        workbook = openpyxl.load_workbook(self.file_path)
        sheet = workbook.active
        for row in sheet.iter_rows(min_row=2, values_only=True):
            salary = json.loads(row[2]) if row[2] else None
            vacancy = Vacancy(
                title=row[0],
                link=row[1],
                salary=salary,
                description=row[3],
                requirements=row[4],
            )
            vacancies.append(vacancy)
        workbook.close()
        return self._filter_vacancies(vacancies, criteria)

    def delete_vacancy(self, criteria: Dict[str, Any]) -> None:
        vacancies = self.get_vacancies({})
        filtered_vacancies = [v for v in vacancies if not self._matches_criteria(v, criteria)]
        self._save_all_vacancies(filtered_vacancies)

    def _save_all_vacancies(self, vacancies: List[Vacancy]) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.append(["title", "link", "salary", "description", "requirements"])
        for vacancy in vacancies:
            sheet.append([
                vacancy.title,
                vacancy.link,
                json.dumps(vacancy.salary) if vacancy.salary else "",
                vacancy.description,
                vacancy.requirements,
            ])
        workbook.save(self.file_path)
        workbook.close()

    def _filter_vacancies(
        self, vacancies: List[Vacancy], criteria: Dict[str, Any]
    ) -> List[Vacancy]:
        filtered = []
        for vacancy in vacancies:
            matches = True
            for key, value in criteria.items():
                if key == "keyword":
                    if value.lower() not in vacancy.description.lower() and value.lower() not in vacancy.requirements.lower():
                        matches = False
                        break
                elif key == "min_salary":
                    if vacancy.get_salary() < value:
                        matches = False
                        break
                elif getattr(vacancy, key, None) != value:
                    matches = False
                    break
            if matches:
                filtered.append(vacancy)
        return filtered

    def _matches_criteria(self, vacancy_data: Dict[str, Any], criteria: Dict[str, Any]) -> bool:
        for key, value in criteria.items():
            if key == "keyword":
                description = vacancy_data.get("description", "").lower()
                requirements = vacancy_data.get("requirements", "").lower()
                title = vacancy_data.get("title", "").lower()
                if (value.lower() not in description and
                        value.lower() not in requirements and
                        value.lower() not in title):
                    return False
            elif key == "min_salary":
                vacancy = Vacancy.validate_and_create(vacancy_data)
                if vacancy.get_salary() < value:
                    return False
            elif getattr(vacancy_data, key, None) != value:
                return False
        return True